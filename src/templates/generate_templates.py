"""
Generate category-wise upload templates for Flowers, Cakes, Plants.
Mandatory fields → yellow background
Optional fields  → white background
Run: python3 generate_templates.py
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ──────────────────────────────────────────────────────────────────
YELLOW     = PatternFill("solid", fgColor="FFD700")   # mandatory
WHITE      = PatternFill("solid", fgColor="FFFFFF")   # optional
HEADER_BG  = PatternFill("solid", fgColor="1F2937")   # dark header row
EXAMPLE_BG = PatternFill("solid", fgColor="F0FFF0")   # light green example row
NOTE_BG    = PatternFill("solid", fgColor="FFF8DC")   # cream note row
WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BLACK_FONT = Font(name="Arial", color="000000", size=9)
RED_FONT   = Font(name="Arial", color="CC0000", bold=True, size=9)
GREY_FONT  = Font(name="Arial", color="555555", italic=True, size=8)

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Shared allowed values ─────────────────────────────────────────────────────
ALL_OCCASION_TAGS = "Anniversary|Appreciation|Baby Shower|Birthday|Cheer Up|Congratulations|Funeral|Get Well Soon|House Warming|I am sorry|Love n Romance|New Born Baby|Thank You|Valentine|Wedding"
ALL_FESTIVAL_TAGS = "New Year|Republic Day|Rose Day|Propose Day|Chocolate Day|Teddy Day|Promise Day|Hug Day|Kiss Day|Valentine's Day|Womens Day|Doctors day|Mothers Day|Brother's Day|Father's Day|Christmas|Lohri|Holi|Eid Al Fitr"
ALL_RELATIONSHIP  = "Her|Him|Friends|Wife|Husband|Kids|Boyfriend|Girlfriend|Mother|Father|Brother|Sister"

# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA DEFINITIONS
# Each column: (header, mandatory, example_value, allowed_values_note)
# ─────────────────────────────────────────────────────────────────────────────

FLOWERS_COLUMNS = [
    # ── Core (Mandatory) ──
    ("name",                               True,  "Red Rose Bouquet",                       "Full product name. Max 200 chars."),
    ("sku",                                True,  "200999FL",                                "Unique. Pattern: XXXXXXFL"),
    ("quantity",                           True,  "50",                                      "Available stock. Integer."),
    ("costing_price",                      True,  "400",                                     "Internal cost price (INR). Number."),
    ("original_price",                     True,  "999",                                     "MRP shown to customer (INR). Number."),
    ("selling_price",                      True,  "849",                                     "Discounted selling price (INR). Number."),
    ("description",                        True,  "A stunning bouquet of 12 fresh red roses wrapped in premium paper.", "Full product description. Max 2000 chars."),
    ("categorization.subcategory_name",    True,  "Roses",                                   "Allowed: Roses | Lillies | Carnation | Roses and Carnation | Roses and Lily"),
    ("categorization.type",                True,  "Hot Pick",                                "Allowed: Bouquet | Hot Pick | Best Seller | New Arrival"),
    ("media.primary_image_url",            True,  "https://firebasestorage.googleapis.com/...", "Primary product image URL. Must start with https://"),
    # ── Optional ──
    ("short_summary",                      False, "12 Red Roses | Premium Wrapping",         "Short product summary. Max 150 chars."),
    ("categorization.occasion_tags",       False, "Birthday|Anniversary|Love n Romance",     f"Pipe-separated (|). Allowed: {ALL_OCCASION_TAGS}"),
    ("categorization.festival_tags",       False, "Valentine's Day|Rose Day|Propose Day",    f"Pipe-separated (|). Allowed: {ALL_FESTIVAL_TAGS}"),
    ("categorization.relationship",        False, "Her|Him|Wife|Girlfriend",                 f"Pipe-separated (|). Allowed: {ALL_RELATIONSHIP}"),
    ("product_attributes.color",           False, "Red",                                     "Primary colour of the flowers."),
    ("product_attributes.product_content", False, "12 Red Roses|Premium Wrapping Paper|Green Fillers", "Pipe-separated list of what's included in the product."),
    ("product_attributes.stem_length_cm",  False, "40",                                      "Stem length in centimetres. Number."),
    ("product_attributes.fragrance_level", False, "Medium",                                  "Allowed: None | Low | Medium | High"),
    ("product_attributes.vase_life_days_min", False, "5",                                   "Minimum vase life in days. Number."),
    ("product_attributes.origin",          False, "Bangalore",                               "Origin city/country of flowers."),
    ("media.gallery_images",               False, "https://cdn.../img2.jpg|https://cdn.../img3.jpg", "Pipe-separated image URLs. Max 5."),
    ("care_and_logistics.care_instructions", False, "Keep in cool place|Change water daily|Trim stems at angle", "Pipe-separated care tips."),
    ("care_and_logistics.requires_cold_chain", False, "FALSE",                               "TRUE or FALSE. Does product need cold storage?"),
    ("care_and_logistics.max_delivery_days",   False, "1",                                   "Max days allowed for delivery. Integer."),
    ("care_and_logistics.regional_availability", False, "Pan India",                         "Delivery regions. Pipe-separated or 'Pan India'."),
    ("availability.is_active",             False, "TRUE",                                    "TRUE = live on site | FALSE = hidden. Default: TRUE"),
    ("availability.is_featured",           False, "FALSE",                                   "TRUE = show in featured section. Default: FALSE"),
]

CAKES_COLUMNS = [
    # ── Core (Mandatory) ──
    ("name",                               True,  "Chocolate Truffle Cake",                  "Full product name. Max 200 chars."),
    ("sku",                                True,  "100999CA",                                "Unique. Pattern: XXXXXXCA"),
    ("quantity",                           True,  "20",                                      "Available stock. Integer."),
    ("costing_price",                      True,  "350",                                     "Internal cost price (INR). Number."),
    ("original_price",                     True,  "754",                                     "MRP shown to customer (INR). Number."),
    ("selling_price",                      True,  "629",                                     "Discounted selling price (INR). Number."),
    ("description",                        True,  "Indulge in layers of rich chocolate sponge and velvety truffle cream.", "Full product description. Max 2000 chars."),
    ("categorization.subcategory_name",    True,  "Chocolate Cakes",                         "Allowed: Black Forest Cakes | Chocolate Cakes | Red Velvet Cakes | Ferrero Rocher Cakes | Pinata cake"),
    ("categorization.type",                True,  "Hot Pick",                                "Allowed: Hot Pick | Best Seller | New Arrival"),
    ("media.primary_image_url",            True,  "https://firebasestorage.googleapis.com/...", "Primary product image URL. Must start with https://"),
    # ── Optional ──
    ("short_summary",                      False, "Rich chocolate sponge | Truffle cream | 500g",  "Short product summary. Max 150 chars."),
    ("categorization.occasion_tags",       False, "Birthday|Anniversary|Congratulations",    f"Pipe-separated (|). Allowed: {ALL_OCCASION_TAGS}"),
    ("categorization.festival_tags",       False, "Christmas|New Year|Valentine's Day",      f"Pipe-separated (|). Allowed: {ALL_FESTIVAL_TAGS}"),
    ("categorization.relationship",        False, "Her|Him|Wife|Husband|Kids|Friends",       f"Pipe-separated (|). Allowed: {ALL_RELATIONSHIP}"),
    ("media.gallery_images",               False, "https://cdn.../img2.jpg|https://cdn.../img3.jpg", "Pipe-separated image URLs. Max 5."),
    ("care_and_logistics.care_instructions", False, "Store in refrigerator|Best consumed within 24 hours|Keep away from heat", "Pipe-separated care tips."),
    ("care_and_logistics.max_delivery_days",   False, "1",                                   "Max days allowed for delivery. Integer."),
    ("availability.is_active",             False, "TRUE",                                    "TRUE = live on site | FALSE = hidden. Default: TRUE"),
    ("availability.is_featured",           False, "FALSE",                                   "TRUE = show in featured section. Default: FALSE"),
    # ── Variations (for different weights/sizes) ──
    ("variations.variant_name",            False, "0.5 Kg",                                  "Variant label shown to customer. E.g. 0.5 Kg | 1 Kg | 1.5 Kg | 2 Kg"),
    ("variations.variant_sku",             False, "100999CA-500G",                           "Unique SKU for this variant."),
    ("variations.weight_kg",               False, "0.5",                                     "Weight of this variant in Kg. Decimal allowed."),
    ("variations.original_price",          False, "754",                                     "MRP for this variant (INR)."),
    ("variations.selling_price",           False, "629",                                     "Selling price for this variant (INR)."),
    ("variations.quantity_available",      False, "15",                                      "Stock available for this variant. Integer."),
]

PLANTS_COLUMNS = [
    # ── Core (Mandatory) ──
    ("name",                               True,  "Laurentii Snake Plant",                   "Full product name. Max 200 chars."),
    ("sku",                                True,  "300999PL",                                "Unique. Pattern: XXXXXXPL"),
    ("quantity",                           True,  "35",                                      "Available stock. Integer."),
    ("costing_price",                      True,  "300",                                     "Internal cost price (INR). Number."),
    ("original_price",                     True,  "594",                                     "MRP shown to customer (INR). Number."),
    ("selling_price",                      True,  "540",                                     "Discounted selling price (INR). Number."),
    ("description",                        True,  "The Laurentii Snake Plant is an air-purifying indoor plant known for its striking variegated leaves.", "Full product description. Max 2000 chars."),
    ("categorization.subcategory_name",    True,  "Snake Plant",                             "Allowed: Snake Plant | Money Plant | Jade Plant | Peace Lily Plant | Syngonium | Bonsai"),
    ("categorization.type",                True,  "Hot Pick",                                "Allowed: Hot Pick | Best Seller | New Arrival"),
    ("media.primary_image_url",            True,  "https://firebasestorage.googleapis.com/...", "Primary product image URL. Must start with https://"),
    # ── Optional ──
    ("short_summary",                      False, "Air Purifying | Low Maintenance | Ceramic Pot", "Short product summary. Max 150 chars."),
    ("categorization.occasion_tags",       False, "Birthday|House Warming|Appreciation",     f"Pipe-separated (|). Allowed: {ALL_OCCASION_TAGS}"),
    ("categorization.festival_tags",       False, "New Year|Mothers Day|Father's Day",       f"Pipe-separated (|). Allowed: {ALL_FESTIVAL_TAGS}"),
    ("categorization.relationship",        False, "Her|Him|Mother|Father|Friends",           f"Pipe-separated (|). Allowed: {ALL_RELATIONSHIP}"),
    ("product_attributes.color",           False, "Green",                                   "Primary colour of plant or pot."),
    ("product_attributes.product_content", False, "1 Snake Plant|1 Ceramic Pot|Care Card",   "Pipe-separated list of what's included."),
    ("product_attributes.origin",          False, "India",                                   "Origin country/region of the plant."),
    ("media.gallery_images",               False, "https://cdn.../img2.jpg|https://cdn.../img3.jpg", "Pipe-separated image URLs. Max 5."),
    ("care_and_logistics.care_instructions", False, "Water once a week|Indirect sunlight|Wipe leaves monthly", "Pipe-separated care tips."),
    ("care_and_logistics.max_delivery_days",   False, "3",                                   "Max days allowed for delivery. Integer."),
    ("care_and_logistics.regional_availability", False, "Pan India",                         "Delivery regions. Pipe-separated or 'Pan India'."),
    ("availability.is_active",             False, "TRUE",                                    "TRUE = live on site | FALSE = hidden. Default: TRUE"),
    ("availability.is_featured",           False, "FALSE",                                   "TRUE = show in featured section. Default: FALSE"),
]

# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_template(columns, category_name, out_path):
    wb = Workbook()

    # ── Sheet 1: Upload Template ──────────────────────────────────────────────
    ws = wb.active
    ws.title = f"{category_name} Upload"
    ws.freeze_panes = "A4"   # freeze header rows

    # Row 1: Legend
    ws.row_dimensions[1].height = 20
    ws["A1"] = f"🟡 Yellow = Mandatory   |   White = Optional   |   Use pipe (|) to separate multiple values   |   Category: {category_name}"
    ws["A1"].font = Font(name="Arial", bold=True, size=9, color="1F2937")
    ws["A1"].fill = PatternFill("solid", fgColor="FFFDE7")
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Row 2: Column Headers
    ws.row_dimensions[2].height = 28
    for col_idx, (header, mandatory, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font   = WHITE_FONT
        cell.fill   = HEADER_BG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Row 3: Mandatory indicator
    ws.row_dimensions[3].height = 18
    for col_idx, (_, mandatory, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value="* REQUIRED" if mandatory else "optional")
        cell.font   = RED_FONT if mandatory else GREY_FONT
        cell.fill   = YELLOW if mandatory else WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Row 4: Example data
    ws.row_dimensions[4].height = 40
    for col_idx, (_, mandatory, example, _) in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=example)
        cell.font = BLACK_FONT
        cell.fill = EXAMPLE_BG
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Rows 5–104: Empty data rows
    for row in range(5, 105):
        ws.row_dimensions[row].height = 20
        for col_idx, (_, mandatory, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value="")
            cell.fill   = YELLOW if mandatory else WHITE
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column widths
    for col_idx, (header, _, _, _) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        if "description" in header or "tags" in header or "instructions" in header or "content" in header:
            ws.column_dimensions[col_letter].width = 45
        elif "image" in header or "gallery" in header:
            ws.column_dimensions[col_letter].width = 40
        elif "name" in header:
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 22

    # ── Sheet 2: Allowed Values Reference ─────────────────────────────────────
    ref = wb.create_sheet("Allowed Values")
    ref.column_dimensions["A"].width = 35
    ref.column_dimensions["B"].width = 10
    ref.column_dimensions["C"].width = 90

    # Header
    for col, title in enumerate(["Column Header", "Required", "Allowed Values / Notes"], start=1):
        cell = ref.cell(row=1, column=col, value=title)
        cell.font  = WHITE_FONT
        cell.fill  = HEADER_BG
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ref.row_dimensions[1].height = 22

    for row_idx, (header, mandatory, _, note) in enumerate(columns, start=2):
        ref.row_dimensions[row_idx].height = 30
        c1 = ref.cell(row=row_idx, column=1, value=header)
        c2 = ref.cell(row=row_idx, column=2, value="YES" if mandatory else "no")
        c3 = ref.cell(row=row_idx, column=3, value=note)
        for cell in [c1, c2, c3]:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        c1.font = Font(name="Arial", bold=True, size=9)
        c2.font = RED_FONT if mandatory else GREY_FONT
        c2.fill = YELLOW if mandatory else WHITE
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c3.font = BLACK_FONT

    wb.save(out_path)
    print(f"✅ {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# GENERATE ALL THREE
# ─────────────────────────────────────────────────────────────────────────────

build_template(FLOWERS_COLUMNS, "Flowers", "/Users/aman/Desktop/Redheart/redheart-backend/src/templates/flowers_upload_template.xlsx")
build_template(CAKES_COLUMNS,   "Cakes",   "/Users/aman/Desktop/Redheart/redheart-backend/src/templates/cakes_upload_template.xlsx")
build_template(PLANTS_COLUMNS,  "Plants",  "/Users/aman/Desktop/Redheart/redheart-backend/src/templates/plants_upload_template.xlsx")

print("\nAll 3 templates generated successfully.")
